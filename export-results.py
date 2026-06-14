"""
export-results.py — Export axe-core scan findings to Excel

Usage:
    SUPABASE_SERVICE_ROLE_KEY=<key> python export-results.py

Get the service role key from:
    Supabase dashboard → Settings → API → service_role (secret)

Output: scan-results-export.xlsx in the current directory
"""

import os
import sys
import json
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL  = "https://vgifjzxnjwieqgltuviv.supabase.co"
SERVICE_KEY   = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SERVICE_KEY:
    print("ERROR: Set SUPABASE_SERVICE_ROLE_KEY env var first.")
    print("  Get it from: Supabase dashboard → Settings → API → service_role")
    print("  Run: SUPABASE_SERVICE_ROLE_KEY=<key> python export-results.py")
    sys.exit(1)

HEADERS = {
    "apikey":        SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type":  "application/json",
}

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", start_color="1E3A5F")
HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT      = Font(name="Arial", size=10)
ALT_FILL       = PatternFill("solid", start_color="F5F7FA")
IMPACT_COLORS  = {
    "critical": "C0392B",
    "serious":  "E67E22",
    "moderate": "F1C40F",
    "minor":    "27AE60",
}

def header_style(cell, text):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── Fetch data ────────────────────────────────────────────────────────────────
print("Fetching audits…")
r = requests.get(
    f"{SUPABASE_URL}/rest/v1/audits",
    headers={**HEADERS, "Prefer": "return=representation"},
    params={"select": "*"},
)
if not r.ok:
    print(f"  ERROR {r.status_code}: {r.text}")
    sys.exit(1)
_audits_raw = r.json()
if _audits_raw:
    print(f"  Columns available: {list(_audits_raw[0].keys())}")
audits = {a["id"]: a for a in _audits_raw}

print("Fetching scan jobs…")
r = requests.get(
    f"{SUPABASE_URL}/rest/v1/scan_jobs",
    headers={**HEADERS, "Prefer": "return=representation"},
    params={"select": "*"},
)
if not r.ok:
    print(f"  ERROR {r.status_code}: {r.text}")
    sys.exit(1)
_jobs_raw = r.json()
if _jobs_raw:
    print(f"  Job columns: {list(_jobs_raw[0].keys())}")
jobs = {j["id"]: j for j in _jobs_raw}

print("Fetching scan results…")
r = requests.get(
    f"{SUPABASE_URL}/rest/v1/scan_results",
    headers={**HEADERS, "Prefer": "return=representation"},
    params={
        "select": "job_id,violations_json,incomplete_json,passes_json,violation_count,incomplete_count,pass_count",
        "order":  "job_id.asc",
    },
)
if not r.ok:
    print(f"  ERROR {r.status_code}: {r.text}")
    sys.exit(1)
results = r.json()
print(f"  Found {len(results)} result records across {len(jobs)} jobs")

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
cols = ["Audit Name", "Website", "WCAG", "Level", "Page / Component",
        "Scan Type", "Status", "Violations", "Incomplete", "Passes",
        "Started At", "Duration (s)"]
for c, h in enumerate(cols, 1):
    header_style(ws.cell(1, c), h)
ws.row_dimensions[1].height = 30

for row_i, res in enumerate(results, 2):
    job   = jobs.get(res["job_id"], {})
    audit = audits.get(job.get("audit_id", ""), {})
    fill  = ALT_FILL if row_i % 2 == 0 else None

    def w(col, val):
        cell = ws.cell(row_i, col, val)
        cell.font = BODY_FONT
        if fill:
            cell.fill = fill

    # Duration
    dur = ""
    if job.get("started_at") and job.get("completed_at"):
        try:
            s = datetime.fromisoformat(job["started_at"].replace("Z", "+00:00"))
            e = datetime.fromisoformat(job["completed_at"].replace("Z", "+00:00"))
            dur = round((e - s).total_seconds())
        except Exception:
            pass

    w(1,  audit.get("name", "—"))
    w(2,  audit.get("website_url", "—"))
    w(3,  audit.get("wcag_version", "—"))
    w(4,  audit.get("conformance_level", "—"))
    w(5,  job.get("scan_name") or job.get("url", "—"))
    w(6,  job.get("scan_type", "—"))
    w(7,  job.get("status", "—"))
    w(8,  res.get("violation_count", 0))
    w(9,  res.get("incomplete_count", 0))
    w(10, res.get("pass_count", 0))
    w(11, (job.get("started_at") or "")[:19].replace("T", " "))
    w(12, dur)

set_widths(ws, [28, 30, 10, 8, 30, 12, 10, 12, 12, 8, 20, 12])
freeze_and_filter(ws)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Violations (one row per violation rule per scan)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Violations")
cols2 = ["Audit Name", "Page / Component", "Scan Type",
         "Rule ID", "Impact", "Description", "Help",
         "WCAG Criteria", "Instances (nodes)", "Help URL"]
for c, h in enumerate(cols2, 1):
    header_style(ws2.cell(1, c), h)
ws2.row_dimensions[1].height = 30

row_i = 2
for res in results:
    job   = jobs.get(res["job_id"], {})
    audit = audits.get(job.get("audit_id", ""), {})
    violations = res.get("violations_json") or []

    for v in violations:
        impact = v.get("impact", "")
        fill_color = IMPACT_COLORS.get(impact)
        impact_fill = PatternFill("solid", start_color=fill_color) if fill_color else None

        wcag_tags = [t for t in v.get("tags", []) if t.startswith("wcag")]
        wcag_str  = ", ".join(wcag_tags)

        def w2(col, val):
            cell = ws2.cell(row_i, col, val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=(col in [6,7,8]))

        w2(1, audit.get("name", "—"))
        w2(2, job.get("scan_name") or job.get("url", "—"))
        w2(3, job.get("scan_type", "—"))
        w2(4, v.get("id", ""))

        # Impact cell — colour coded
        imp_cell = ws2.cell(row_i, 5, impact)
        imp_cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        if impact_fill:
            imp_cell.fill = impact_fill
        imp_cell.alignment = Alignment(horizontal="center")

        w2(6,  v.get("description", ""))
        w2(7,  v.get("help", ""))
        w2(8,  wcag_str)
        w2(9,  len(v.get("nodes", [])))

        url_cell = ws2.cell(row_i, 10, v.get("helpUrl", ""))
        url_cell.font      = Font(name="Arial", size=10, color="1155CC", underline="single")
        url_cell.alignment = Alignment(wrap_text=False)

        row_i += 1

set_widths(ws2, [28, 30, 12, 28, 10, 45, 35, 22, 10, 45])
freeze_and_filter(ws2)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Violation Nodes (one row per affected DOM element)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Nodes (all instances)")
cols3 = ["Audit Name", "Page / Component", "Rule ID", "Impact",
         "Failure Summary", "HTML Snippet", "Target Selector", "Help URL"]
for c, h in enumerate(cols3, 1):
    header_style(ws3.cell(1, c), h)
ws3.row_dimensions[1].height = 30

row_i = 2
for res in results:
    job   = jobs.get(res["job_id"], {})
    audit = audits.get(job.get("audit_id", ""), {})
    violations = res.get("violations_json") or []

    for v in violations:
        impact = v.get("impact", "")
        fill_color = IMPACT_COLORS.get(impact)

        for node in v.get("nodes", []):
            alt = row_i % 2 == 0

            def w3(col, val):
                cell = ws3.cell(row_i, col, val)
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=(col in [5,6,7]))
                if alt and col not in [4]:
                    cell.fill = ALT_FILL

            w3(1, audit.get("name", "—"))
            w3(2, job.get("scan_name") or job.get("url", "—"))
            w3(3, v.get("id", ""))

            imp_cell = ws3.cell(row_i, 4, impact)
            imp_cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            if fill_color:
                imp_cell.fill = PatternFill("solid", start_color=fill_color)
            imp_cell.alignment = Alignment(horizontal="center")

            # Failure summary — flatten any/all checks
            checks = (node.get("any") or []) + (node.get("all") or [])
            summary = " | ".join(
                c.get("message", "") for c in checks if c.get("message")
            )[:500]

            w3(5, summary)
            w3(6, (node.get("html") or "")[:300])
            w3(7, str(node.get("target") or ""))

            url_cell = ws3.cell(row_i, 8, v.get("helpUrl", ""))
            url_cell.font      = Font(name="Arial", size=10, color="1155CC", underline="single")
            url_cell.alignment = Alignment()

            row_i += 1

set_widths(ws3, [28, 30, 28, 10, 50, 40, 30, 45])
freeze_and_filter(ws3)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Incomplete (needs review)
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Incomplete (review)")
cols4 = ["Audit Name", "Page / Component", "Rule ID", "Impact",
         "Description", "WCAG Criteria", "Instances", "Help URL"]
for c, h in enumerate(cols4, 1):
    header_style(ws4.cell(1, c), h)
ws4.row_dimensions[1].height = 30

row_i = 2
for res in results:
    job   = jobs.get(res["job_id"], {})
    audit = audits.get(job.get("audit_id", ""), {})
    incomplete = res.get("incomplete_json") or []

    for v in incomplete:
        impact = v.get("impact", "")
        wcag_tags = [t for t in v.get("tags", []) if t.startswith("wcag")]
        alt = row_i % 2 == 0

        def w4(col, val):
            cell = ws4.cell(row_i, col, val)
            cell.font = BODY_FONT
            if alt and col != 4:
                cell.fill = ALT_FILL

        w4(1, audit.get("name", "—"))
        w4(2, job.get("scan_name") or job.get("url", "—"))
        w4(3, v.get("id", ""))

        imp_cell = ws4.cell(row_i, 4, impact)
        imp_cell.font = Font(bold=True, name="Arial", size=10,
                             color="FFFFFF" if impact in IMPACT_COLORS else "000000")
        if impact in IMPACT_COLORS:
            imp_cell.fill = PatternFill("solid", start_color=IMPACT_COLORS[impact])
        imp_cell.alignment = Alignment(horizontal="center")

        w4(5, v.get("description", ""))
        w4(6, ", ".join(wcag_tags))
        w4(7, len(v.get("nodes", [])))

        url_cell = ws4.cell(row_i, 8, v.get("helpUrl", ""))
        url_cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")
        row_i += 1

set_widths(ws4, [28, 30, 28, 10, 45, 22, 10, 45])
freeze_and_filter(ws4)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "scan-results-export.xlsx"
wb.save(out)
print(f"\n✓ Saved: {out}")
print(f"  Sheets: Summary | Violations | Nodes (all instances) | Incomplete (review)")

# Quick stats
total_violations = sum(r.get("violation_count", 0) for r in results)
total_incomplete = sum(r.get("incomplete_count", 0) for r in results)
print(f"\n  Total violations: {total_violations}")
print(f"  Total incomplete: {total_incomplete}")
print(f"  Scans exported:   {len(results)}")
